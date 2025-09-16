
from openpyxl import load_workbook
class ExcelUtils:

    @staticmethod
    def get_excel_Data(file_path,sheet_name):
        wb=load_workbook(file_path)
        sh=wb[sheet_name]
        data=[]

        for row in sh.iter_rows(min_row=2,values_only=True):
            data.append(row)

        return data

